import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === Known 2nd aorist root verbs ===
second_aorist_roots = {
    'ἄγω', 'αἱρέω', 'ἁμαρτάνω', 'ἀποθνῄσκω', 'ἄρχω',
    'βαίνω', 'βάλλω', 'γίνομαι', 'γινώσκω', 'δίδωμι',
    'ἐγείρω', 'εἶδον', 'ἐσθίω', 'εὑρίσκω', 'ἔχω',
    'ἔρχομαι', 'ἵστημι', 'ἵημι', 'λαγχάνω', 'λαμβάνω',
    'λέγω', 'λείπω', 'μανθάνω', 'ὁράω', 'πάσχω',
    'πείθω', 'πίνω', 'πίπτω', 'πυνθάνομαι', 'στρέφω',
    'τίθημι', 'τρέχω', 'τυγχάνω', 'φαίνω', 'φεύγω',
    'φέρω', 'φημί', 'χαίρω', 'ἀνοίγω', 'ἀποστέλλω',
    'ἐπιτρέπω', 'σπείρω', 'τρέφω',
}

# 2nd aorist stem for display
stem_display = {
    'ἄγω': 'ἀγαγ-',
    'αἱρέω': 'ἑλ-',
    'ἁμαρτάνω': 'ἁμαρτ-',
    'ἀνοίγω': 'ἠνοιγ- (2aor pass)',
    'ἀποθνῄσκω': 'θαν-',
    'ἀποστέλλω': 'σταλ- (2aor pass)',
    'ἄρχω': 'ἀρξ- (mid)',
    'βαίνω': '-βη-/-βα-',
    'βάλλω': 'βαλ-',
    'γίνομαι': 'γεν-',
    'γινώσκω': 'γνω-/γνο-',
    'δίδωμι': 'δο-',
    'εἶδον': 'ἰδ-',
    'ἐσθίω': 'φαγ-',
    'εὑρίσκω': 'εὑρ-',
    'ἔρχομαι': 'ἐλθ-',
    'ἔχω': 'σχ-',
    'ἵημι': 'ἑ-',
    'ἵστημι': 'στη-/στα-',
    'λαγχάνω': 'λαχ-',
    'λαμβάνω': 'λαβ-',
    'λέγω': 'εἰπ-',
    'λείπω': 'λιπ-',
    'μανθάνω': 'μαθ-',
    'ὁράω': 'ἰδ-',
    'πάσχω': 'παθ-',
    'πείθω': 'πιθ- (2aor pass)',
    'πίνω': 'πι-',
    'πίπτω': 'πεσ-/πεσο-',
    'πυνθάνομαι': 'πυθ-',
    'στρέφω': 'στραφ- (2aor pass)',
    'τίθημι': 'θε-/θη-',
    'τρέχω': 'δραμ-',
    'τυγχάνω': 'τυχ-',
    'φαίνω': 'φαν- (2aor pass)',
    'φεύγω': 'φυγ-',
    'φέρω': 'ἐνεγκ-',
    'φημί': 'ἐφη-',
    'χαίρω': 'χαρ- (2aor pass)',
    'σπείρω': 'σπαρ- (2aor pass)',
    'τρέφω': 'τραφ- (2aor pass)',
    'ἐγείρω': 'ἐγερ- (2aor pass)',
    'ἐπιτρέπω': 'τραπ- (2aor pass)',
}

# Compound verb -> root mapping
# We map known compound lemmas to their root
compound_to_root = {
    # ἄγω compounds
    'ἀνάγω': 'ἄγω', 'ἀπάγω': 'ἄγω', 'εἰσάγω': 'ἄγω', 'ἐξάγω': 'ἄγω',
    'ἐπάγω': 'ἄγω', 'κατάγω': 'ἄγω', 'παράγω': 'ἄγω', 'προάγω': 'ἄγω',
    'προσάγω': 'ἄγω', 'συνάγω': 'ἄγω', 'ὑπάγω': 'ἄγω',
    # αἱρέω compounds
    'ἀναιρέω': 'αἱρέω', 'ἀφαιρέω': 'αἱρέω', 'ἐξαιρέω': 'αἱρέω',
    'καθαιρέω': 'αἱρέω', 'περιαιρέω': 'αἱρέω', 'διαιρέω': 'αἱρέω',
    # βαίνω compounds
    'ἀναβαίνω': 'βαίνω', 'ἀποβαίνω': 'βαίνω', 'διαβαίνω': 'βαίνω',
    'ἐμβαίνω': 'βαίνω', 'ἐπιβαίνω': 'βαίνω', 'καταβαίνω': 'βαίνω',
    'μεταβαίνω': 'βαίνω', 'παραβαίνω': 'βαίνω', 'προβαίνω': 'βαίνω',
    'συγκαταβαίνω': 'βαίνω', 'συμβαίνω': 'βαίνω', 'συναναβαίνω': 'βαίνω',
    'ὑπερβαίνω': 'βαίνω',
    # βάλλω compounds
    'ἀναβάλλω': 'βάλλω', 'ἀποβάλλω': 'βάλλω', 'ἐκβάλλω': 'βάλλω',
    'ἐμβάλλω': 'βάλλω', 'ἐπιβάλλω': 'βάλλω', 'καταβάλλω': 'βάλλω',
    'μεταβάλλω': 'βάλλω', 'παραβάλλω': 'βάλλω', 'περιβάλλω': 'βάλλω',
    'προβάλλω': 'βάλλω', 'συμβάλλω': 'βάλλω', 'ὑποβάλλω': 'βάλλω',
    # γίνομαι compounds
    'διαγίνομαι': 'γίνομαι', 'ἐπιγίνομαι': 'γίνομαι', 'παραγίνομαι': 'γίνομαι',
    # γινώσκω compounds
    'ἀναγινώσκω': 'γινώσκω', 'ἐπιγινώσκω': 'γινώσκω', 'καταγινώσκω': 'γινώσκω',
    # δίδωμι compounds
    'ἀναδίδωμι': 'δίδωμι', 'ἀποδίδωμι': 'δίδωμι', 'ἐπιδίδωμι': 'δίδωμι',
    'μεταδίδωμι': 'δίδωμι', 'παραδίδωμι': 'δίδωμι', 'προδίδωμι': 'δίδωμι',
    # ἔρχομαι compounds
    'ἀπέρχομαι': 'ἔρχομαι', 'διέρχομαι': 'ἔρχομαι', 'εἰσέρχομαι': 'ἔρχομαι',
    'ἐξέρχομαι': 'ἔρχομαι', 'ἐπέρχομαι': 'ἔρχομαι', 'κατέρχομαι': 'ἔρχομαι',
    'παρέρχομαι': 'ἔρχομαι', 'περιέρχομαι': 'ἔρχομαι', 'προέρχομαι': 'ἔρχομαι',
    'προσέρχομαι': 'ἔρχομαι', 'συνέρχομαι': 'ἔρχομαι', 'ὑπέρχομαι': 'ἔρχομαι',
    'ἀντιπαρέρχομαι': 'ἔρχομαι',
    # ἵστημι compounds
    'ἀνίστημι': 'ἵστημι', 'ἀφίστημι': 'ἵστημι', 'ἀνθίστημι': 'ἵστημι',
    'ἐνίστημι': 'ἵστημι', 'ἐξανίστημι': 'ἵστημι', 'ἐξίστημι': 'ἵστημι',
    'ἐπίστημι': 'ἵστημι', 'ἐφίστημι': 'ἵστημι', 'καθίστημι': 'ἵστημι',
    'κατεφίστημι': 'ἵστημι', 'κατεφιστάμαι': 'ἵστημι',
    'μεθίστημι': 'ἵστημι', 'παρίστημι': 'ἵστημι', 'περιίστημι': 'ἵστημι',
    'προΐστημι': 'ἵστημι', 'συνίστημι': 'ἵστημι', 'συνεφίστημι': 'ἵστημι',
    'κατεπίσταμαι': 'ἵστημι',
    # ἵημι compounds
    'ἀφίημι': 'ἵημι', 'ἀνίημι': 'ἵημι', 'συνίημι': 'ἵημι',
    # λαμβάνω compounds
    'ἀναλαμβάνω': 'λαμβάνω', 'ἀντιλαμβάνω': 'λαμβάνω',
    'ἐπιλαμβάνω': 'λαμβάνω', 'καταλαμβάνω': 'λαμβάνω',
    'μεταλαμβάνω': 'λαμβάνω', 'παραλαμβάνω': 'λαμβάνω',
    'προσλαμβάνω': 'λαμβάνω', 'συλλαμβάνω': 'λαμβάνω',
    'συμπαραλαμβάνω': 'λαμβάνω', 'συμπεριλαμβάνω': 'λαμβάνω',
    'ὑπολαμβάνω': 'λαμβάνω',
    # λέγω compounds
    'ἀντιλέγω': 'λέγω', 'διαλέγομαι': 'λέγω', 'ἐκλέγομαι': 'λέγω',
    'ἐπιλέγω': 'λέγω', 'καταλέγω': 'λέγω', 'προλέγω': 'λέγω',
    # λείπω compounds
    'ἀπολείπω': 'λείπω', 'ἐκλείπω': 'λείπω', 'ἐπιλείπω': 'λείπω',
    'καταλείπω': 'λείπω', 'περιλείπω': 'λείπω', 'ὑπολείπω': 'λείπω',
    # τίθημι compounds
    'ἀνατίθημι': 'τίθημι', 'ἀποτίθημι': 'τίθημι', 'διατίθημι': 'τίθημι',
    'ἐκτίθημι': 'τίθημι', 'ἐπιτίθημι': 'τίθημι', 'κατατίθημι': 'τίθημι',
    'μετατίθημι': 'τίθημι', 'παρατίθημι': 'τίθημι', 'περιτίθημι': 'τίθημι',
    'προστίθημι': 'τίθημι', 'συντίθημι': 'τίθημι', 'ὑποτίθημι': 'τίθημι',
    'συνεπιτίθεμαι': 'τίθημι',
    # πίπτω compounds
    'ἐκπίπτω': 'πίπτω', 'ἐπιπίπτω': 'πίπτω', 'καταπίπτω': 'πίπτω',
    'παραπίπτω': 'πίπτω', 'περιπίπτω': 'πίπτω', 'προσπίπτω': 'πίπτω',
    # φεύγω compounds
    'ἀποφεύγω': 'φεύγω', 'διαφεύγω': 'φεύγω', 'ἐκφεύγω': 'φεύγω',
    'καταφεύγω': 'φεύγω',
    # τρέχω compounds
    'εἰστρέχω': 'τρέχω', 'ἐπισυντρέχω': 'τρέχω', 'κατατρέχω': 'τρέχω',
    'περιτρέχω': 'τρέχω', 'προστρέχω': 'τρέχω', 'συντρέχω': 'τρέχω',
    'ὑποτρέχω': 'τρέχω',
    # ὁράω compounds
    'ἀφοράω': 'ὁράω', 'ἐφοράω': 'ὁράω', 'καθοράω': 'ὁράω',
    'προοράω': 'ὁράω', 'συνοράω': 'ὁράω', 'ὑπεροράω': 'ὁράω',
    # εὑρίσκω compounds
    'ἀνευρίσκω': 'εὑρίσκω',
    # τυγχάνω compounds
    'ἐντυγχάνω': 'τυγχάνω', 'ἐπιτυγχάνω': 'τυγχάνω',
    'παρατυγχάνω': 'τυγχάνω', 'ὑπερεντυγχάνω': 'τυγχάνω',
    # ἔχω compounds
    'ἀνέχω': 'ἔχω', 'ἀντέχω': 'ἔχω', 'ἀπέχω': 'ἔχω',
    'ἐνέχω': 'ἔχω', 'ἐπέχω': 'ἔχω', 'κατέχω': 'ἔχω',
    'μετέχω': 'ἔχω', 'παρέχω': 'ἔχω', 'περιέχω': 'ἔχω',
    'προέχω': 'ἔχω', 'προσέχω': 'ἔχω', 'συνέχω': 'ἔχω',
    'ὑπέχω': 'ἔχω', 'ὑπερέχω': 'ἔχω',
    # στρέφω compounds
    'ἀναστρέφω': 'στρέφω', 'ἀποστρέφω': 'στρέφω', 'διαστρέφω': 'στρέφω',
    'ἐπιστρέφω': 'στρέφω', 'καταστρέφω': 'στρέφω', 'μεταστρέφω': 'στρέφω',
    'ὑποστρέφω': 'στρέφω',
    # ἀποστέλλω compounds
    'ἐξαποστέλλω': 'ἀποστέλλω', 'συναποστέλλω': 'ἀποστέλλω',
    # σπείρω compounds
    'διασπείρω': 'σπείρω', 'ἐπισπείρω': 'σπείρω',
    # τρέφω compounds
    'ἀνατρέφω': 'τρέφω', 'ἐκτρέφω': 'τρέφω', 'ἐντρέφω': 'τρέφω',
    # φαίνω compounds
    'ἀναφαίνω': 'φαίνω', 'ἐπιφαίνω': 'φαίνω',
    # ἐσθίω compounds
    'κατεσθίω': 'ἐσθίω', 'συνεσθίω': 'ἐσθίω',
    # πίνω compounds
    'συμπίνω': 'πίνω',
    # πάσχω - no common compounds
    # ἀποθνῄσκω - compound of θνῄσκω really
    'συναποθνῄσκω': 'ἀποθνῄσκω',
    # περιτέμνω - has 2nd aor? Actually no, it is 1st aor
    # ἐγείρω compounds
    'διεγείρω': 'ἐγείρω', 'ἐξεγείρω': 'ἐγείρω', 'συνεγείρω': 'ἐγείρω',
    # εἶδον is suppletive aorist of ὁράω - merge with ὁράω
    'εἶδον': 'ὁράω',
    # κατανύσσω - not really a standard 2nd aorist root
    # φέρω compounds
    'ἀναφέρω': 'φέρω', 'ἀποφέρω': 'φέρω', 'διαφέρω': 'φέρω',
    'εἰσφέρω': 'φέρω', 'ἐκφέρω': 'φέρω', 'ἐπιφέρω': 'φέρω',
    'καταφέρω': 'φέρω', 'παραφέρω': 'φέρω', 'περιφέρω': 'φέρω',
    'προσφέρω': 'φέρω', 'συμφέρω': 'φέρω', 'ὑποφέρω': 'φέρω',
}

# Also add self-mappings for roots
for r in second_aorist_roots:
    if r not in compound_to_root:
        compound_to_root[r] = r

def get_root(lemma):
    if lemma in compound_to_root:
        return compound_to_root[lemma]
    return None

# Parse voice/mood from MorphGNT parsing code
def parse_info(parsing):
    voices = {'A': 'Active', 'M': 'Middle', 'P': 'Passive'}
    moods = {'I': 'Indicative', 'S': 'Subjunctive', 'O': 'Optative',
             'D': 'Imperative', 'N': 'Infinitive', 'P': 'Participle'}
    v = voices.get(parsing[2], '?') if len(parsing) > 2 else '?'
    m = moods.get(parsing[3], '?') if len(parsing) > 3 else '?'
    return v, m

# Read MorphGNT Acts
with open('morphgnt/65-Ac-morphgnt.txt', 'r', encoding='utf-8') as f:
    morphlines = f.readlines()

# Extract all aorist verbs and check against 2nd aorist roots
root_groups = defaultdict(lambda: {
    'lemmas': set(), 'forms': set(), 'count': 0,
    'voice_moods': set(), 'refs': set()
})

for line in morphlines:
    parts = line.strip().split()
    if len(parts) < 7:
        continue
    ref, pos, parsing, text, word, norm, lemma = (
        parts[0], parts[1], parts[2], parts[3], parts[4], parts[5], parts[6]
    )
    if pos != 'V-' or len(parsing) < 4 or parsing[1] != 'A':
        continue  # not an aorist verb

    root = get_root(lemma)
    if root is None:
        continue

    v, m = parse_info(parsing)
    ch = int(ref[2:4])
    vs = int(ref[4:6])

    root_groups[root]['lemmas'].add(lemma)
    root_groups[root]['forms'].add(norm)
    root_groups[root]['count'] += 1
    root_groups[root]['voice_moods'].add(f'{v} {m}')
    root_groups[root]['refs'].add(f'{ch}:{vs}')

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "2nd Aorist by Root"

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
greek_font = Font(name='SBL BibLit', size=11)
greek_font_bold = Font(name='SBL BibLit', size=11, bold=True)
normal_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')

headers = ['#', 'Root Verb', '2nd Aorist Stem', 'Compounds in Acts',
           'Voice/Mood', 'Unique Forms', 'Occurrences', 'Sample References']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

row = 2
for idx, root in enumerate(sorted(root_groups.keys()), 1):
    data = root_groups[root]
    compounds = sorted(data['lemmas'] - {root})
    compound_str = ', '.join(compounds) if compounds else '(none)'
    vm_str = ', '.join(sorted(data['voice_moods']))
    forms_str = ', '.join(sorted(data['forms']))
    refs = sorted(data['refs'], key=lambda r: (int(r.split(':')[0]), int(r.split(':')[1])))
    ref_sample = ', '.join(refs[:8])
    if len(refs) > 8:
        ref_sample += f'  ... ({len(refs)} total)'

    values = [
        idx,
        root,
        stem_display.get(root, ''),
        compound_str,
        vm_str,
        forms_str,
        data['count'],
        ref_sample,
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        if col in (2, 3, 4, 6):  # Greek columns
            cell.font = greek_font_bold if col == 2 else greek_font
        else:
            cell.font = normal_font
    row += 1

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 35

# Freeze top row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:H{row-1}'

wb.save('acts-2nd-aorist-by-root.xlsx')
print(f'Done: {row-2} root verbs written to acts-2nd-aorist-by-root.xlsx')
